"""Render the season aggregates into the Hebrew dashboard and the Excel workbook.

The dashboard is a single self-contained RTL HTML file with its data inlined, so it
opens from disk with no server and no network.

Chart colours come from the dataviz reference palette unchanged - categorical slots
1-3 (blue / orange / aqua), which that palette documents as validated on the all-pairs
gate in both light and dark modes. The club red is used for chrome and headings only,
never as a series colour, so it can never be mistaken for an encoded value.
"""
from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sources.common import DATA, OUTPUT, RESOURCES, load_config, read_json, setup_logging

LOG = setup_logging("build")

GOALKEEPER = "שוער"


def labels() -> dict:
    return read_json(RESOURCES / "labels_he.json", default={}) or {}


# --------------------------------------------------------------------------- data

def collect(cfg: dict) -> dict:
    season = read_json(DATA / f"season_{cfg['season']['slug']}.json")
    if not season:
        raise SystemExit("no season aggregate found - run hta_aggregate.py first")

    schedule = read_json(DATA / "schedule_status.json", default={}) or {}
    delta = read_json(DATA / "last_delta.json", default={}) or {}
    fixtures = (read_json(DATA / "fixtures.json", default={}) or {}).get("fixtures", [])

    # The CI build deliberately skips hta_schedule.py - that exists only to drive the
    # Windows task - so there is no schedule_status.json there. Fall back to the fixture
    # list so the "next match" tile still appears on the published page.
    next_match = schedule.get("next_match")
    if not next_match and fixtures:
        nxt = next((f for f in fixtures if f.get("start_time")), None)
        if nxt:
            local = datetime.fromisoformat(nxt["start_time"].replace("Z", "+00:00")).astimezone(
                ZoneInfo(cfg["timezone"])
            )
            next_match = {
                "game_id": nxt.get("game_id"),
                "competition": nxt.get("competition_name"),
                "home": nxt.get("home"),
                "away": nxt.get("away"),
                "kickoff_local": local.strftime("%Y-%m-%d %H:%M"),
            }

    # Normalised to exactly the fields the page renders. The scheduler emits a richer
    # object than the fixture fallback does, and carrying that extra baggage made the
    # PC-built page differ from the CI-built one for no visible reason - which showed up
    # as two different fingerprints for what is really the same page.
    schedule = dict(schedule)
    schedule["next_match"] = (
        {
            "game_id": next_match.get("game_id"),
            "competition": next_match.get("competition"),
            "home": next_match.get("home"),
            "away": next_match.get("away"),
            "kickoff_local": next_match.get("kickoff_local"),
        }
        if next_match
        else None
    )
    fetch_status = read_json(DATA / "fetch_status.json", default={}) or {}

    matches = []
    for path in (DATA / "matches").glob("*.json"):
        m = read_json(path)
        if m:
            matches.append(m)
    matches.sort(key=lambda m: m.get("start_time") or "", reverse=True)

    # The source's "missing" list is captured in each match file but is NOT surfaced.
    # It proved untrustworthy: 8 of 10 matches reported nobody missing, return dates
    # were years stale ("Late October 2024"), and three players it listed as injured
    # went on to play every subsequent match. Better to show nothing than something
    # confidently wrong.

    record = {"w": 0, "d": 0, "l": 0, "gf": 0, "ga": 0, "cs": 0, "matches": 0}
    for comp in season.get("competitions", []):
        record["w"] += comp["w"]
        record["d"] += comp["d"]
        record["l"] += comp["l"]
        record["gf"] += comp["goals_for"]
        record["ga"] += comp["goals_against"]
        record["cs"] += comp["clean_sheets"]
        record["matches"] += comp["matches"]

    return {
        "season": season,
        "schedule": schedule,
        "delta": delta,
        "fixtures": fixtures[:8],
        "matches": [
            {
                "game_id": m["game_id"],
                "date": (m.get("start_time") or "")[:10],
                "competition": m.get("competition_name"),
                "competition_id": m.get("competition_id"),
                "opponent": m.get("opponent"),
                "is_home": m.get("is_home"),
                "team_score": m.get("team_score"),
                "opponent_score": m.get("opponent_score"),
                "result": m.get("result"),
                "clean_sheet": m.get("team_clean_sheet"),
                "stats_complete": m.get("stats_complete", True),
            }
            for m in matches
        ],
        "record": record,
        "source": fetch_status.get("source", "365scores"),
        "last_run": read_json(DATA / "last_run.json", default={}) or {},
        "generated_at": datetime.now(ZoneInfo(cfg["timezone"])).strftime("%Y-%m-%d %H:%M"),
        "labels": labels(),
        "competition_names": cfg["competitions"]["names_he"],
        "color_slots": cfg["competitions"].get("color_slots", {}),
    }


# --------------------------------------------------------------------------- excel

HEADER_FILL = PatternFill("solid", fgColor="C8102E")
HEADER_FONT = Font(bold=True, color="FFFFFF")

SHEET_COLUMNS = [
    ("name", "player"), ("position", "position"), ("jersey", "jersey"),
    ("apps", "apps"), ("starts", "starts"), ("bench_apps", "bench_apps"),
    ("subbed_off", "subbed_off"), ("minutes", "minutes"), ("goals", "goals"),
    ("assists", "assists"), ("goal_involvements", "goal_involvements"),
    ("clean_sheets", "clean_sheets"), ("yellow", "yellow"),
    ("second_yellow", "second_yellow"), ("red", "red"),
    ("avg_rating", "avg_rating"), ("minutes_per_goal", "minutes_per_goal"),
]


def _style_sheet(ws, ncols: int, nrows: int, widths=None) -> None:
    ws.sheet_view.rightToLeft = True
    for idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = (widths or {}).get(idx, 14)
    ws.freeze_panes = "A2"
    if nrows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{nrows}"


def _player_sheet(wb, title: str, rows: list[dict], lab: dict) -> None:
    ws = wb.create_sheet(title[:31])
    ws.append([lab.get(key, key) for _, key in SHEET_COLUMNS])
    for row in rows:
        ws.append([row.get(field) for field, _ in SHEET_COLUMNS])

    ncols, nrows = len(SHEET_COLUMNS), len(rows) + 1
    _style_sheet(ws, ncols, nrows, widths={1: 22, 2: 10})

    if nrows > 1:
        minutes_col = get_column_letter([f for f, _ in SHEET_COLUMNS].index("minutes") + 1)
        ws.conditional_formatting.add(
            f"{minutes_col}2:{minutes_col}{nrows}",
            ColorScaleRule(
                start_type="min", start_color="FFF4E6",
                end_type="max", end_color="C8102E",
            ),
        )


def build_excel(ctx: dict, path: Path) -> None:
    lab, season = ctx["labels"], ctx["season"]
    wb = Workbook()
    wb.remove(wb.active)

    _player_sheet(wb, lab.get("all_competitions", "סה\"כ"), season["total"], lab)

    for comp in season.get("competitions", []):
        rows = season["by_competition"].get(str(comp["id"]), [])
        if rows:
            _player_sheet(wb, comp["name"], rows, lab)

    # Goalkeepers get their own sheet - saves and concessions are meaningless elsewhere.
    gk_rows = [r for r in season["total"] if r.get("position") == GOALKEEPER]
    if gk_rows:
        ws = wb.create_sheet(lab.get("goalkeepers", "שוערים")[:31])
        ws.append([lab.get(k, k) for k in
                   ["player", "apps", "starts", "minutes", "clean_sheets",
                    "goals_conceded", "saves", "save_pct", "avg_rating"]])
        for row in gk_rows:
            faced = (row["saves"] or 0) + (row["goals_conceded"] or 0)
            ws.append([
                row["name"], row["apps"], row["starts"], row["minutes"],
                row["clean_sheets"], row["goals_conceded"], row["saves"],
                round(100 * row["saves"] / faced, 1) if faced else None,
                row["avg_rating"],
            ])
        _style_sheet(ws, 9, len(gk_rows) + 1, widths={1: 22})

    # Fixtures and results.
    ws = wb.create_sheet(lab.get("fixtures", "משחקים")[:31])
    ws.append([lab.get(k, k) for k in
               ["date", "competition", "opponent", "home", "result", "clean_sheets"]])
    for m in reversed(ctx["matches"]):
        ws.append([
            m["date"], m["competition"], m["opponent"],
            lab.get("home", "בית") if m["is_home"] else lab.get("away", "חוץ"),
            f"{m['team_score']}-{m['opponent_score']}",
            "כן" if m["clean_sheet"] else "לא",
        ])
    for f in ctx["fixtures"]:
        ws.append([
            (f.get("start_time") or "")[:10], f.get("competition_name"),
            f"{f.get('home')} - {f.get('away')}", "", "", "",
        ])
    _style_sheet(ws, 6, len(ctx["matches"]) + len(ctx["fixtures"]) + 1, widths={2: 18, 3: 26})

    # Changelog.
    changelog = DATA / "changelog.md"
    if changelog.exists():
        ws = wb.create_sheet(lab.get("changelog", "יומן")[:31])
        ws.append([lab.get("changelog", "יומן עדכונים")])
        for line in changelog.read_text(encoding="utf-8").splitlines():
            ws.append([line])
        _style_sheet(ws, 1, 1, widths={1: 90})

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    LOG.info("wrote %s (%d sheets)", path.name, len(wb.sheetnames))


# --------------------------------------------------------------------------- html

TEMPLATE = """<!doctype html>
<html lang="he" dir="rtl">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<meta name="hta-build" content="__FINGERPRINT__">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Assistant:wght@400;600;700&display=swap">
<style>
:root {
  color-scheme: light;
  --surface-1: #fcfcfb;
  --page: #f9f9f7;
  --text-primary: #0b0b0b;
  --text-secondary: #52514e;
  --muted: #898781;
  --grid: #e1e0d9;
  --axis: #c3c2b7;
  --border: rgba(11,11,11,0.10);
  --series-1: #2a78d6;
  --series-2: #eb6834;
  --series-3: #1baf7a;
  --seq-200: #9ec5f4;
  --seq-450: #2a78d6;
  --good: #0ca30c;
  --warning: #fab219;
  --critical: #d03b3b;
  --club: #c8102e;
  --club-ink: #c8102e;
}
@media (prefers-color-scheme: dark) {
  :root:not([data-theme="light"]) {
    color-scheme: dark;
    --surface-1: #1a1a19;
    --page: #0d0d0d;
    --text-primary: #ffffff;
    --text-secondary: #c3c2b7;
    --muted: #898781;
    --grid: #2c2c2a;
    --axis: #383835;
    --border: rgba(255,255,255,0.10);
    --series-1: #3987e5;
    --series-2: #d95926;
    --series-3: #199e70;
    --seq-200: #1c5cab;
    --seq-450: #3987e5;
    --club-ink: #ff6b7f;
  }
}
:root[data-theme="dark"] {
  color-scheme: dark;
  --surface-1: #1a1a19;
  --page: #0d0d0d;
  --text-primary: #ffffff;
  --text-secondary: #c3c2b7;
  --muted: #898781;
  --grid: #2c2c2a;
  --axis: #383835;
  --border: rgba(255,255,255,0.10);
  --series-1: #3987e5;
  --series-2: #d95926;
  --series-3: #199e70;
  --seq-200: #1c5cab;
  --seq-450: #3987e5;
  --club-ink: #ff6b7f;
}

* { box-sizing: border-box; }
/* An author `display` rule (.banner is display:flex) beats the UA default for [hidden],
   so el.hidden alone would not hide it. */
[hidden] { display: none !important; }
body {
  margin: 0; padding: 0 0 48px;
  background: var(--page);
  color: var(--text-primary);
  /* Assistant is a Hebrew-first UI face; the stack falls back cleanly offline. */
  font: 14px/1.55 "Assistant", system-ui, -apple-system, "Segoe UI", sans-serif;
}
.wrap { max-width: 1240px; margin: 0 auto; padding: 0 16px; }

header.top {
  background: var(--surface-1);
  border-bottom: 3px solid var(--club);
  padding: 18px 0 14px;
  margin-bottom: 16px;
}
.title-row { display: flex; align-items: baseline; gap: 12px; flex-wrap: wrap; }
h1 { margin: 0; font-size: 22px; letter-spacing: -0.01em; color: var(--club-ink); }
.sub { color: var(--text-secondary); font-size: 13px; }
.theme-btn {
  background: transparent; color: var(--text-secondary);
  border: 1px solid var(--border); border-radius: 8px; padding: 5px 11px;
  cursor: pointer; font: inherit; font-size: 12px;
}
#opsBtn { margin-inline-start: auto; }
#opsBtn[aria-pressed="true"] { color: var(--text-primary); border-color: var(--series-1); }

.tiles { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 10px; margin-top: 14px; }
.tile { background: var(--surface-1); border: 1px solid var(--border); border-radius: 10px; padding: 10px 12px; }
.tile .k { font-size: 11px; color: var(--muted); }
.tile .v { font-size: 21px; font-weight: 650; margin-top: 2px; }
.tile .n { font-size: 11px; color: var(--text-secondary); }

.banner {
  border-radius: 10px; padding: 10px 13px; margin: 12px 0;
  border: 1px solid var(--border); background: var(--surface-1);
  display: flex; gap: 9px; align-items: flex-start; font-size: 13px;
}
.banner .icon { font-size: 15px; line-height: 1.2; }
.banner.warn { border-inline-start: 4px solid var(--warning); }
.banner.alert { border-inline-start: 4px solid var(--critical); }
.banner.info { border-inline-start: 4px solid var(--series-1); }

section.card {
  background: var(--surface-1); border: 1px solid var(--border);
  border-radius: 12px; padding: 14px 16px; margin-bottom: 16px;
}
section.card > h2 { margin: 0 0 10px; font-size: 15px; font-weight: 650; }
.card-head { display: flex; align-items: center; gap: 10px; margin-bottom: 10px; }
.card-head h2 { margin: 0; font-size: 15px; font-weight: 650; }
.mini-sel {
  margin-inline-start: auto; font: inherit; font-size: 12px; padding: 4px 8px;
  border-radius: 8px; border: 1px solid var(--border);
  background: var(--page); color: var(--text-secondary); cursor: pointer;
}
.mini-sel:focus-visible, .tab:focus-visible, input.search:focus-visible,
.theme-btn:focus-visible { outline: 2px solid var(--series-1); outline-offset: 2px; }

.controls { display: flex; gap: 8px; flex-wrap: wrap; align-items: center; margin-bottom: 12px; }
.tabs { display: flex; gap: 6px; flex-wrap: wrap; }
.tab {
  border: 1px solid var(--border); background: transparent; color: var(--text-secondary);
  border-radius: 999px; padding: 6px 13px; cursor: pointer; font: inherit; font-size: 13px;
}
.tab[aria-selected="true"] { background: var(--club); border-color: var(--club); color: #fff; font-weight: 600; }
input.search {
  margin-inline-start: auto; padding: 7px 11px; border-radius: 8px;
  border: 1px solid var(--border); background: var(--page); color: var(--text-primary);
  font: inherit; font-size: 13px; min-width: 180px;
}

.tbl-scroll { overflow-x: auto; }
/* Sticky headers need a scroll container with a bounded height. Without the
   max-height the header sticks to the top of a box that is itself taller than the
   viewport, so it scrolls away with the page and appears not to work at all. */
.tbl-scroll.tall { max-height: 72vh; overflow: auto; }
table { width: 100%; border-collapse: separate; border-spacing: 0; font-variant-numeric: tabular-nums; }
th, td { padding: 7px 8px; text-align: center; white-space: nowrap; }
th {
  font-size: 11px; color: var(--muted); font-weight: 600;
  border-bottom: 1px solid var(--axis); cursor: pointer; user-select: none;
  position: sticky; top: 0; z-index: 2; background: var(--surface-1);
}
th:hover { color: var(--text-primary); }
th.name-col, td.name-col { text-align: right; }
/* Keep the player's name in view when scrolling the columns sideways. */
th.name-col, td.name-col { position: sticky; right: 0; background: var(--surface-1); }
td.name-col { z-index: 1; }
th.name-col { z-index: 3; }
tbody tr { border-bottom: 1px solid var(--grid); }
tbody td { border-bottom: 1px solid var(--grid); }
tbody tr:hover td { background: color-mix(in srgb, var(--series-1) 7%, var(--surface-1)); cursor: pointer; }
td.name-col { font-weight: 600; }
/* Opaque, not translucent - these rows pass underneath the sticky header. */
.pos-row td { background: color-mix(in srgb, var(--muted) 14%, var(--surface-1));
  font-size: 11px; color: var(--text-secondary); font-weight: 650; text-align: right; }
.pos-row:hover td { background: color-mix(in srgb, var(--muted) 14%, var(--surface-1)); cursor: default; }
.jersey { color: var(--muted); font-weight: 400; font-size: 11px; margin-inline-start: 5px; }
.dim { color: var(--muted); }

.minbar { position: relative; min-width: 86px; }
.minbar .track { height: 5px; background: var(--grid); border-radius: 3px; overflow: hidden; margin-top: 3px; }
.minbar .fill { height: 100%; background: var(--seq-450); border-radius: 3px; }

.chart-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 16px; }
.bars { display: flex; flex-direction: column; gap: 7px; }
.bar-row { display: grid; grid-template-columns: 108px 1fr 34px; gap: 9px; align-items: center; font-size: 12px; }
.bar-row .lbl { text-align: left; color: var(--text-secondary); overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.bar-track { display: block; background: var(--grid); border-radius: 4px; height: 15px; overflow: hidden; }
/* display:block matters - an inline span ignores a percentage width and the bar vanishes. */
.bar-fill { display: block; height: 100%; border-radius: 4px; min-width: 2px; }
.bar-row .val { font-weight: 650; font-variant-numeric: tabular-nums; }
.legend { display: flex; gap: 14px; flex-wrap: wrap; font-size: 12px; color: var(--text-secondary); margin-bottom: 9px; }
.legend .sw { display: inline-block; width: 10px; height: 10px; border-radius: 3px; margin-inline-end: 5px; }

ul.plain { list-style: none; margin: 0; padding: 0; }
ul.plain li { padding: 5px 0; border-bottom: 1px solid var(--grid); font-size: 13px; }
ul.plain li:last-child { border-bottom: 0; }
.chip { display: inline-block; padding: 1px 7px; border-radius: 999px; font-size: 11px;
  background: color-mix(in srgb, var(--series-1) 15%, transparent); margin-inline-end: 5px; }
.chip.good { background: color-mix(in srgb, var(--good) 20%, transparent); }
.chip.bad { background: color-mix(in srgb, var(--critical) 20%, transparent); }

dialog.modal {
  border: 1px solid var(--border); border-radius: 14px; padding: 0;
  background: var(--surface-1); color: var(--text-primary);
  max-width: 760px; width: 94vw; max-height: 84vh;
}
dialog.modal::backdrop { background: rgba(0,0,0,0.45); }
.modal-head { display: flex; align-items: center; gap: 10px; padding: 14px 16px; border-bottom: 1px solid var(--grid); }
.modal-head h3 { margin: 0; font-size: 16px; }
.modal-body { padding: 12px 16px 18px; overflow-y: auto; max-height: 66vh; }
.close-btn { margin-inline-start: auto; background: transparent; border: 1px solid var(--border);
  border-radius: 8px; color: var(--text-secondary); cursor: pointer; padding: 4px 10px; font: inherit; }
footer.foot { color: var(--muted); font-size: 12px; text-align: center; margin-top: 22px; }
.res-W { color: var(--good); font-weight: 650; }
.res-L { color: var(--critical); font-weight: 650; }
.res-D { color: var(--text-secondary); font-weight: 650; }
</style>
</head>
<body>
<header class="top"><div class="wrap">
  <div class="title-row">
    <h1>__TITLE__</h1>
    <span class="sub" id="stamp" title="">__SEASON__ · __UPDATED__</span>
    <button class="theme-btn" id="opsBtn" type="button" aria-pressed="false" hidden>מידע תחזוקה</button>
    <button class="theme-btn" id="themeBtn" type="button">מצב תצוגה</button>
  </div>
  <div class="tiles" id="tiles"></div>
</div></header>

<div class="wrap">
  <div id="banners"></div>

  <section class="card">
    <h2>__L_WHAT_CHANGED__</h2>
    <div id="delta"></div>
  </section>

  <section class="card">
    <div class="controls">
      <div class="tabs" id="tabs" role="tablist"></div>
      <input class="search" id="search" type="search" placeholder="__L_SEARCH__" aria-label="__L_SEARCH__">
    </div>
    <div class="tbl-scroll tall"><table id="mainTable">
      <thead><tr id="headRow"></tr></thead>
      <tbody id="tableBody"></tbody>
    </table></div>
    <p class="sub" style="margin:10px 0 0">לחיצה על שורת שחקן פותחת פירוט משחק אחר משחק.</p>
  </section>

  <section class="card">
    <h2>__L_GK__</h2>
    <div class="tbl-scroll"><table id="gkTable">
      <thead><tr id="gkHead"></tr></thead>
      <tbody id="gkBody"></tbody>
    </table></div>
  </section>

  <div class="chart-grid">
    <section class="card">
      <div class="card-head">
        <h2>__L_SCORERS__</h2>
        <select class="mini-sel" id="scorersComp" aria-label="__L_COMP__"></select>
      </div>
      <div class="bars" id="scorers"></div>
    </section>
    <section class="card">
      <div class="card-head">
        <h2>__L_ASSISTS__</h2>
        <select class="mini-sel" id="assistsComp" aria-label="__L_COMP__"></select>
      </div>
      <div class="bars" id="assisters"></div>
    </section>
  </div>

  <section class="card">
    <h2>__L_MINCOMP__</h2>
    <div class="legend" id="compLegend"></div>
    <div class="bars" id="compBars"></div>
  </section>

  <section class="card">
    <h2>__L_MATCHES__</h2>
    <div class="tbl-scroll tall"><table>
      <thead><tr><th>__L_DATE__</th><th>__L_COMP__</th><th>__L_OPP__</th><th></th>
        <th>__L_RESULT__</th><th>__L_CS__</th></tr></thead>
      <tbody id="matchBody"></tbody>
    </table></div>
  </section>

  <footer class="foot ops">__L_SOURCE__: __SOURCE__ · __UPDATED__</footer>
</div>

<dialog class="modal" id="playerModal">
  <div class="modal-head">
    <h3 id="modalName"></h3>
    <button class="close-btn" id="modalClose" type="button">סגירה</button>
  </div>
  <div class="modal-body" id="modalBody"></div>
</dialog>

<script id="payload" type="application/json">__DATA__</script>
<script>
const DATA = JSON.parse(document.getElementById('payload').textContent);
const L = DATA.labels;
const SERIES = ['--series-1','--series-2','--series-3'];

/* ---------- theme toggle (persisted per viewer) ---------- */
(function () {
  let saved = null;
  try { saved = localStorage.getItem('hta-theme'); } catch (e) {}
  if (saved) document.documentElement.setAttribute('data-theme', saved);
  document.getElementById('themeBtn').addEventListener('click', () => {
    const cur = document.documentElement.getAttribute('data-theme');
    const isDark = cur ? cur === 'dark'
      : window.matchMedia('(prefers-color-scheme: dark)').matches;
    const next = isDark ? 'light' : 'dark';
    document.documentElement.setAttribute('data-theme', next);
    try { localStorage.setItem('hta-theme', next); } catch (e) {}
  });
})();

const esc = s => String(s == null ? '' : s)
  .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');

/* ---------- maintenance detail, hidden by default ----------
   Elements marked .ops are for whoever maintains this page, not for someone the link
   was shared with. Hidden by default so a shared link is clean with no special URL —
   query strings are not something to rely on surviving the artifact wrapper. Warning
   and alert banners are deliberately NOT .ops: those are data-correctness signals that
   every reader should see. */
function applyOps(show) {
  document.querySelectorAll('.ops').forEach(el => { el.hidden = !show; });
  const btn = document.getElementById('opsBtn');
  if (btn) btn.setAttribute('aria-pressed', show ? 'true' : 'false');
}

let opsShown = false;
let opsKnown = false;
try {
  const saved = localStorage.getItem('hta-ops');
  opsKnown = saved !== null;
  opsShown = saved === '1';
} catch (e) {}

const opsBtn = document.getElementById('opsBtn');
// The button itself is hidden from anyone the link was shared with - seeing a
// "maintenance info" control on someone else's dashboard is just confusing. It appears
// only for a viewer who has opted in on this device before, or after the reveal gesture
// below. Three taps on the header stamp works on a phone as well as a desktop.
if (opsKnown) opsBtn.hidden = false;

opsBtn.addEventListener('click', () => {
  opsShown = !opsShown;
  try { localStorage.setItem('hta-ops', opsShown ? '1' : '0'); } catch (e) {}
  applyOps(opsShown);
});

let taps = 0, tapTimer = null;
document.getElementById('stamp').addEventListener('click', () => {
  taps += 1;
  clearTimeout(tapTimer);
  tapTimer = setTimeout(() => { taps = 0; }, 700);
  if (taps >= 3) {
    taps = 0;
    opsBtn.hidden = false;
    try { localStorage.setItem('hta-ops', opsShown ? '1' : '0'); } catch (e) {}
  }
});

/* ---------- header tiles ---------- */
(function () {
  const r = DATA.record;
  const tiles = [
    // Spelled out rather than "4-4-2", which is ambiguous beside an RTL formation.
    { k: L.matches, v: r.matches,
      n: `${r.w} ${L.wins} · ${r.d} ${L.draws} · ${r.l} ${L.losses}` },
    { k: L.goals, v: r.gf, n: `${L.goals_conceded} ${r.ga}` },
    { k: L.clean_sheets, v: r.cs, n: '' },
  ];
  const nm = DATA.schedule && DATA.schedule.next_match;
  if (nm) tiles.push({ k: L.next_match, v: nm.kickoff_local.slice(5).replace('-','/'),
                       n: `${nm.home} - ${nm.away}` });
  document.getElementById('tiles').innerHTML = tiles.map(t =>
    `<div class="tile"><div class="k">${esc(t.k)}</div><div class="v">${esc(t.v)}</div>
     <div class="n">${esc(t.n)}</div></div>`).join('');
})();

/* ---------- banners: schedule flag, fallback source, missing stats ---------- */
(function () {
  const out = [];
  const s = DATA.schedule || {};
  if (s.flag_he) {
    const cls = s.severity === 'alert' ? 'alert' : 'warn';
    const icon = s.severity === 'alert' ? '\\u26D4' : '\\u26A0\\uFE0F';
    out.push(`<div class="banner ${cls}"><span class="icon">${icon}</span><div>
      <strong>${esc(s.flag_he)}</strong><br>
      <span class="sub">הבדיקה הבאה: ${esc(s.next_run_local)}</span></div></div>`);
  } else if (s.next_run_local) {
    // Owner-only ('ops'): operational detail, hidden in the shared view. The header's
    // "עודכן" stamp stays visible in both views, so the staleness signal is never lost.
    const lr = DATA.last_run || {};
    const ran = lr.ran_at ? ` · ריצה אחרונה: <strong>${esc(lr.ran_at)}</strong>` +
                            (lr.ok === false ? ' <span class="chip bad">נכשלה</span>' : '') : '';
    out.push(`<div class="banner info ops"><span class="icon">\\u2713</span><div>
      לוח המשחקים מעודכן. הבדיקה הבאה: <strong>${esc(s.next_run_local)}</strong>${ran}</div></div>`);
  }
  if (DATA.source && DATA.source !== '365scores') {
    out.push(`<div class="banner warn"><span class="icon">\\u26A0\\uFE0F</span><div>
      המקור הראשי לא היה זמין. הנתונים נלקחו מ־<strong>${esc(DATA.source)}</strong>
      ועשויים להיות חלקיים.</div></div>`);
  }
  const miss = DATA.season.matches_missing_stats || [];
  const chips = list => list.map(m =>
    `<span class="chip">${esc(m.date)} ${esc(m.competition)} · ${esc(m.opponent)}</span>`).join('');
  // Rebuilt-from-events and genuinely-absent are different claims; keep them apart.
  const derived = miss.filter(m => m.derived);
  const absent = miss.filter(m => !m.derived);
  if (derived.length) {
    // Owner-only: data-provenance detail. The stats it describes are shown regardless.
    out.push(`<div class="banner info ops"><span class="icon">\\u2139\\uFE0F</span><div>
      ${esc(L.derived_note)}<br>${chips(derived)}</div></div>`);
  }
  if (absent.length) {
    out.push(`<div class="banner warn"><span class="icon">\\u26A0\\uFE0F</span><div>
      ${esc(L.incomplete_note)}<br>${chips(absent)}</div></div>`);
  }
  document.getElementById('banners').innerHTML = out.join('');
})();

/* ---------- what changed ---------- */
(function () {
  const d = DATA.delta || {};
  const el = document.getElementById('delta');
  if (d.is_first_run) {
    el.innerHTML = `<p class="sub">טעינה ראשונית — ${esc(DATA.season.matches_counted)} משחקים נטענו.</p>`;
    return;
  }
  if (!d.changes || !d.changes.length) {
    el.innerHTML = `<p class="sub">${esc(L.no_changes)}</p>`;
    return;
  }
  el.innerHTML = '<ul class="plain">' + d.changes.map(c => {
    const parts = Object.entries(c.diff).map(([k, v]) => {
      const cls = (k === 'yellow' || k === 'red' || k === 'second_yellow') ? 'bad' : 'good';
      return `<span class="chip ${cls}">${v > 0 ? '+' : ''}${v} ${esc(L[k] || k)}</span>`;
    }).join('');
    return `<li><strong>${esc(c.name)}</strong> ${parts}</li>`;
  }).join('') + '</ul>';
})();

/* ---------- main table ---------- */
const COLS = [
  { f: 'name', l: L.player, name: true },
  { f: 'apps', l: L.apps },
  { f: 'starts', l: L.starts },
  { f: 'bench_apps', l: L.bench_apps },
  { f: 'minutes', l: L.minutes, bar: true },
  { f: 'goals', l: L.goals },
  { f: 'assists', l: L.assists },
  { f: 'clean_sheets', l: L.clean_sheets },
  { f: 'yellow', l: L.yellow },
  { f: 'second_yellow', l: L.second_yellow },
  { f: 'red', l: L.red },
  { f: 'avg_rating', l: L.avg_rating },
  { f: 'minutes_per_goal', l: L.minutes_per_goal },
];

let activeComp = 'total';
let sortField = 'minutes';
let sortDir = -1;
let query = '';

function rowsFor(comp) {
  return comp === 'total' ? DATA.season.total : (DATA.season.by_competition[comp] || []);
}

function renderHead() {
  document.getElementById('headRow').innerHTML = COLS.map(c => {
    const arrow = sortField === c.f ? (sortDir === -1 ? ' \\u25BC' : ' \\u25B2') : '';
    return `<th data-f="${c.f}" class="${c.name ? 'name-col' : ''}">${esc(c.l)}${arrow}</th>`;
  }).join('');
  document.querySelectorAll('#headRow th').forEach(th => {
    th.addEventListener('click', () => {
      const f = th.dataset.f;
      if (sortField === f) sortDir = -sortDir;
      else { sortField = f; sortDir = f === 'name' ? 1 : -1; }
      renderHead(); renderBody();
    });
  });
}

function renderBody() {
  let rows = rowsFor(activeComp).slice();
  if (query) rows = rows.filter(r => r.name.includes(query));

  const maxMin = Math.max(1, ...rows.map(r => r.minutes || 0));
  rows.sort((a, b) => {
    const x = a[sortField], y = b[sortField];
    if (x == null && y == null) return 0;
    if (x == null) return 1;
    if (y == null) return -1;
    if (typeof x === 'string') return x.localeCompare(y, 'he') * sortDir;
    return (x - y) * sortDir;
  });

  // Position grouping only makes sense while sorted by a non-name field.
  const order = L.positions_order || [];
  const groups = new Map();
  rows.forEach(r => {
    const p = r.position || '—';
    if (!groups.has(p)) groups.set(p, []);
    groups.get(p).push(r);
  });
  const ordered = [...groups.keys()].sort((a, b) => {
    const ia = order.indexOf(a), ib = order.indexOf(b);
    return (ia === -1 ? 99 : ia) - (ib === -1 ? 99 : ib);
  });

  let html = '';
  const groupNames = L.position_groups || {};
  ordered.forEach(pos => {
    // Section headings name the unit (הגנה), not the individual position (מגן).
    const heading = groupNames[pos] || pos;
    html += `<tr class="pos-row"><td colspan="${COLS.length}">${esc(heading)}</td></tr>`;
    groups.get(pos).forEach(r => {
      html += `<tr data-pid="${r.player_id}">` + COLS.map(c => {
        const v = r[c.f];
        if (c.name) {
          return `<td class="name-col">${esc(v)}` +
            (r.jersey != null ? `<span class="jersey">${esc(r.jersey)}</span>` : '') + `</td>`;
        }
        if (c.bar) {
          const pct = Math.round(100 * (v || 0) / maxMin);
          return `<td class="minbar">${v == null ? '' : esc(v)}
            <div class="track"><div class="fill" style="width:${pct}%"></div></div></td>`;
        }
        if (v == null || v === 0) return `<td class="dim">${v == null ? '–' : 0}</td>`;
        return `<td>${esc(v)}</td>`;
      }).join('') + `</tr>`;
    });
  });
  document.getElementById('tableBody').innerHTML = html;

  document.querySelectorAll('#tableBody tr[data-pid]').forEach(tr => {
    tr.addEventListener('click', () => openPlayer(tr.dataset.pid));
  });
}

function renderTabs() {
  const tabs = [{ id: 'total', name: L.all_competitions }].concat(
    (DATA.season.competitions || []).map(c => ({ id: String(c.id), name: c.name })));
  document.getElementById('tabs').innerHTML = tabs.map(t =>
    `<button class="tab" role="tab" data-id="${t.id}"
      aria-selected="${t.id === activeComp}">${esc(t.name)}</button>`).join('');
  document.querySelectorAll('#tabs .tab').forEach(b => {
    b.addEventListener('click', () => {
      activeComp = b.dataset.id;
      try { localStorage.setItem('hta-comp', activeComp); } catch (e) {}
      renderTabs(); renderBody();
    });
  });
}

document.getElementById('search').addEventListener('input', e => {
  query = e.target.value.trim();
  renderBody();
});

/* ---------- goalkeepers ---------- */
(function () {
  const keys = ['player','apps','starts','minutes','clean_sheets','goals_conceded','saves','save_pct','avg_rating'];
  document.getElementById('gkHead').innerHTML =
    keys.map((k, i) => `<th class="${i === 0 ? 'name-col' : ''}">${esc(L[k] || k)}</th>`).join('');
  // Only keepers who actually played - the rest would be a wall of zeroes.
  const gks = DATA.season.total.filter(r => r.position === '\\u05E9\\u05D5\\u05E2\\u05E8' && r.apps > 0);
  document.getElementById('gkBody').innerHTML = gks.map(r => {
    const faced = (r.saves || 0) + (r.goals_conceded || 0);
    const pct = faced ? Math.round(1000 * r.saves / faced) / 10 : null;
    return `<tr><td class="name-col">${esc(r.name)}</td><td>${r.apps}</td><td>${r.starts}</td>
      <td>${r.minutes}</td><td>${r.clean_sheets}</td><td>${r.goals_conceded}</td>
      <td>${r.saves}</td><td>${pct == null ? '–' : pct + '%'}</td>
      <td>${r.avg_rating == null ? '–' : r.avg_rating}</td></tr>`;
  }).join('');
})();

/* ---------- bar charts ---------- */
function barChart(el, rows, field, colorVar) {
  const data = rows.filter(r => (r[field] || 0) > 0)
    .sort((a, b) => b[field] - a[field]).slice(0, 8);
  if (!data.length) { el.innerHTML = '<p class="sub">אין נתונים עדיין.</p>'; return; }
  const max = data[0][field];
  el.innerHTML = data.map(r => `
    <div class="bar-row">
      <span class="lbl">${esc(r.name)}</span>
      <span class="bar-track"><span class="bar-fill"
        style="width:${Math.round(100 * r[field] / max)}%;background:var(${colorVar})"></span></span>
      <span class="val">${r[field]}</span>
    </div>`).join('');
}

/* Each chart filters independently of the roster table's tab, so you can look at
   league scorers without changing what the table below is showing. */
function competitionOptions() {
  return [{ id: 'total', name: L.all_competitions }].concat(
    (DATA.season.competitions || []).map(c => ({ id: String(c.id), name: c.name })));
}

function wireChart(selectId, targetId, field, storageKey) {
  const sel = document.getElementById(selectId);
  const opts = competitionOptions();
  sel.innerHTML = opts.map(o => `<option value="${esc(o.id)}">${esc(o.name)}</option>`).join('');

  let saved = null;
  try { saved = localStorage.getItem(storageKey); } catch (e) {}
  if (saved && opts.some(o => o.id === saved)) sel.value = saved;

  const draw = () => barChart(document.getElementById(targetId),
                              rowsFor(sel.value), field, '--seq-450');
  sel.addEventListener('change', () => {
    try { localStorage.setItem(storageKey, sel.value); } catch (e) {}
    draw();
  });
  draw();
}

wireChart('scorersComp', 'scorers', 'goals', 'hta-scorers-comp');
wireChart('assistsComp', 'assisters', 'assists', 'hta-assists-comp');

/* ---------- minutes by competition (identity -> categorical slots 1-3) ---------- */
(function () {
  const comps = DATA.season.competitions || [];
  // Colour follows the competition itself, from a fixed map, so sorting by minutes or
  // a new cup appearing never repaints the others.
  const slots = DATA.color_slots || {};
  let nextFree = Object.keys(slots).length + 1;
  const totals = comps.map(c => {
    const rows = DATA.season.by_competition[String(c.id)] || [];
    const slotNo = slots[String(c.id)] || nextFree++;
    return { name: c.name, matches: c.matches,
             minutes: rows.reduce((s, r) => s + (r.minutes || 0), 0),
             slot: SERIES[(slotNo - 1) % SERIES.length] };
  }).filter(t => t.minutes > 0).sort((a, b) => b.minutes - a.minutes);

  document.getElementById('compLegend').innerHTML = totals.map(t =>
    `<span><span class="sw" style="background:var(${t.slot})"></span>${esc(t.name)}</span>`).join('');

  const max = Math.max(1, ...totals.map(t => t.minutes));
  document.getElementById('compBars').innerHTML = totals.map(t => `
    <div class="bar-row">
      <span class="lbl">${esc(t.name)}</span>
      <span class="bar-track"><span class="bar-fill"
        style="width:${Math.round(100 * t.minutes / max)}%;background:var(${t.slot})"></span></span>
      <span class="val">${t.minutes}</span>
    </div>`).join('');
})();

/* ---------- matches ---------- */
(function () {
  document.getElementById('matchBody').innerHTML = (DATA.matches || []).map(m => `
    <tr><td>${esc(m.date)}</td><td>${esc(m.competition)}</td><td>${esc(m.opponent)}</td>
    <td class="dim">${m.is_home ? esc(L.home) : esc(L.away)}</td>
    <td class="res-${esc(m.result)}">${esc(m.team_score)}-${esc(m.opponent_score)}</td>
    <td>${m.clean_sheet ? '\\u2713' : ''}</td></tr>`).join('');
})();

/* ---------- player timeline modal ---------- */
const modal = document.getElementById('playerModal');
document.getElementById('modalClose').addEventListener('click', () => modal.close());

function openPlayer(pid) {
  const row = DATA.season.total.find(r => String(r.player_id) === String(pid));
  const games = DATA.season.timeline[String(pid)] || [];
  if (!row) return;
  document.getElementById('modalName').textContent = row.name;
  const head = `<tr><th>${esc(L.date)}</th><th>${esc(L.competition)}</th><th>${esc(L.opponent)}</th>
    <th>${esc(L.result)}</th><th></th><th>${esc(L.minutes)}</th><th>${esc(L.goals)}</th>
    <th>${esc(L.assists)}</th><th>${esc(L.clean_sheets)}</th><th>${esc(L.avg_rating)}</th></tr>`;
  const body = games.slice().reverse().map(g => `
    <tr><td>${esc(g.date)}</td><td>${esc(g.competition)}</td><td>${esc(g.opponent)}</td>
      <td class="res-${esc(g.result)}">${esc(g.score)}</td>
      <td class="dim">${g.started ? 'הרכב' : 'ספסל'}</td>
      <td>${esc(g.minutes)}</td><td>${g.goals || ''}</td><td>${g.assists || ''}</td>
      <td>${g.clean_sheet ? '\\u2713' : ''}</td>
      <td>${g.rating == null ? '' : esc(g.rating)}</td></tr>`).join('');
  document.getElementById('modalBody').innerHTML =
    `<div class="tbl-scroll"><table><thead>${head}</thead><tbody>${body}</tbody></table></div>`;
  modal.showModal();
}

/* ---------- init ---------- */
try {
  const savedComp = localStorage.getItem('hta-comp');
  if (savedComp && (savedComp === 'total' || DATA.season.by_competition[savedComp])) activeComp = savedComp;
} catch (e) {}
renderTabs();
renderHead();
renderBody();
// Must run after the banners are built, since .ops elements are created dynamically.
applyOps(opsShown);
</script>
</body>
</html>
"""


def build_fingerprint(payload: dict) -> str:
    """A hash of what the page SAYS, deliberately not of when it was built.

    The cloud routine uses this to decide whether the artifact needs republishing. A
    naive "do the two pages differ?" check is useless here, because the header carries a
    build timestamp that changes on every single build - it would always report a
    difference and degrade into republishing ten near-identical versions a day.

    So generated_at and last_run are excluded, and the template is included: a code or
    layout change must move the fingerprint even when the data is untouched, which is
    exactly the case that silently drifted before.
    """
    season = payload.get("season") or {}
    schedule = payload.get("schedule") or {}
    material = {
        # season.generated_at moves on every aggregate run, so it is dropped; everything
        # else in the aggregate is real content.
        "season": {k: v for k, v in season.items() if k != "generated_at"},
        # Only the next fixture is shown from the schedule. computed_at / next_run_* tick
        # on their own and would make the fingerprint meaningless.
        "next_match": schedule.get("next_match"),
        "delta": payload.get("delta"),
        "matches": payload.get("matches"),
        "record": payload.get("record"),
        "source": payload.get("source"),
        "labels": payload.get("labels"),
        "color_slots": payload.get("color_slots"),
        "template": TEMPLATE,
    }
    blob = json.dumps(material, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()[:12]


def build_html(ctx: dict, path: Path) -> None:
    lab = ctx["labels"]
    payload = {
        "season": ctx["season"],
        "schedule": ctx["schedule"],
        "delta": ctx["delta"],
        "matches": ctx["matches"],
        "record": ctx["record"],
        "source": ctx["source"],
        "labels": lab,
        "color_slots": ctx["color_slots"],
        "last_run": ctx["last_run"],
    }
    replacements = {
        "__FINGERPRINT__": build_fingerprint(payload),
        "__TITLE__": lab.get("title", "הפועל תל אביב"),
        "__SEASON__": ctx["season"].get("season", ""),
        "__UPDATED__": f"{lab.get('last_updated', 'עודכן')} {ctx['generated_at']}",
        "__SOURCE__": ctx["source"],
        # "</" would end the enclosing <script> block early if it ever appeared in data.
        "__DATA__": json.dumps(payload, ensure_ascii=False).replace("</", "<\\/"),
        "__L_WHAT_CHANGED__": lab.get("what_changed", ""),
        "__L_SEARCH__": lab.get("search", ""),
        "__L_GK__": lab.get("goalkeepers", ""),
        "__L_SCORERS__": lab.get("top_scorers", ""),
        "__L_ASSISTS__": lab.get("top_assists", ""),
        "__L_MINCOMP__": lab.get("minutes_by_comp", ""),
        "__L_MATCHES__": lab.get("fixtures", ""),
        "__L_DATE__": lab.get("date", ""),
        "__L_COMP__": lab.get("competition", ""),
        "__L_OPP__": lab.get("opponent", ""),
        "__L_RESULT__": lab.get("result", ""),
        "__L_CS__": lab.get("clean_sheets", ""),
        "__L_SOURCE__": lab.get("source", ""),
    }
    html = TEMPLATE
    for key, value in replacements.items():
        html = html.replace(key, str(value))

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(html, encoding="utf-8")
    LOG.info("wrote %s (%.0f KB)", path.name, path.stat().st_size / 1024)


def build_artifact(source: Path, path: Path) -> None:
    """Emit a body-only variant for publishing as an Artifact.

    The Artifact host supplies its own <!doctype>/<html>/<head>/<body> skeleton, so the
    page must ship as content only. That also means dir="rtl" cannot live on <html>, so
    it moves onto a wrapper element and onto body via CSS.
    """
    full = source.read_text(encoding="utf-8")

    head_start = full.index("<title>")
    head_end = full.index("</head>")
    head = full[head_start:head_end].strip()

    body_start = full.index(">", full.index("<body")) + 1
    body_end = full.rindex("</body>")
    body = full[body_start:body_end].strip()

    rtl_css = (
        "<style>\n"
        "/* The host owns <html>/<body>, so direction is set here rather than on <html>. */\n"
        "body { direction: rtl; unicode-bidi: isolate; }\n"
        "</style>"
    )
    path.write_text(
        f"{head}\n{rtl_css}\n<div dir=\"rtl\" lang=\"he\">\n{body}\n</div>\n",
        encoding="utf-8",
    )
    LOG.info("wrote %s (%.0f KB)", path.name, path.stat().st_size / 1024)


def main() -> int:
    cfg = load_config()
    ctx = collect(cfg)
    build_html(ctx, OUTPUT / "hta_dashboard.html")
    build_artifact(OUTPUT / "hta_dashboard.html", OUTPUT / "hta_dashboard_artifact.html")
    build_excel(ctx, OUTPUT / "hta_stats.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
